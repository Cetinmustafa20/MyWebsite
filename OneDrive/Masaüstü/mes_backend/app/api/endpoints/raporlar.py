from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List, Optional
from datetime import datetime, date
from pydantic import BaseModel
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from app.db.database import get_db
from app.models.models import MakineIsEmri, Ariza, Operator, Kullanici, Makine, KaliteKontrol
from app.core.security import get_current_user, require_yonetici

router = APIRouter(prefix="/raporlar", tags=["raporlar"])


def _dk_fmt(dk):
    if not dk:
        return "-"
    s = int(dk) // 60
    d = int(dk) % 60
    return f"{s}s {d}dk" if s > 0 else f"{d}dk"


def _excel_header_style():
    fill = PatternFill("solid", fgColor="1E3A5F")
    font = Font(bold=True, color="FFFFFF", size=11)
    alignment = Alignment(horizontal="center", vertical="center")
    return fill, font, alignment


# ── JSON RAPORLAR ──────────────────────────────────────────

class OperatorVerimlilik(BaseModel):
    operator_id: int
    ad_soyad: str
    sicil_no: str
    tamamlanan_is: int
    toplam_sure_dakika: Optional[int]
    ort_sure_dakika: Optional[float]


class MakineGecmis(BaseModel):
    makine_is_emri_id: int
    is_emri_no: str
    stok_adi: str
    operator_adi: str
    baslangic: Optional[datetime]
    bitis: Optional[datetime]
    sure_dakika: Optional[int]
    durum: str


@router.get("/operator-verimlilik", response_model=List[OperatorVerimlilik])
def operator_verimlilik(
    baslangic: Optional[date] = Query(None),
    bitis: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    _: Kullanici = Depends(get_current_user),
):
    q = (
        db.query(Operator, func.count(MakineIsEmri.id).label("tamamlanan"))
        .join(MakineIsEmri, MakineIsEmri.operator_id == Operator.id)
        .filter(MakineIsEmri.durum == "tamamlandi")
    )
    if baslangic:
        q = q.filter(MakineIsEmri.baslangic_zamani >= datetime.combine(baslangic, datetime.min.time()))
    if bitis:
        q = q.filter(MakineIsEmri.bitis_zamani <= datetime.combine(bitis, datetime.max.time()))
    q = q.group_by(Operator.id, Operator.sicil_no, Operator.kullanici_id, Operator.departman, Operator.aktif, Operator.olusturma_tarihi)

    result = []
    for op, tamamlanan in q.all():
        mie_list = db.query(MakineIsEmri).filter(
            MakineIsEmri.operator_id == op.id,
            MakineIsEmri.durum == "tamamlandi",
            MakineIsEmri.baslangic_zamani != None,
            MakineIsEmri.bitis_zamani != None,
        ).all()
        sureler = [
            int((m.bitis_zamani - m.baslangic_zamani).total_seconds() / 60)
            for m in mie_list if m.baslangic_zamani and m.bitis_zamani
        ]
        toplam = sum(sureler) if sureler else None
        ort = round(sum(sureler) / len(sureler), 1) if sureler else None
        result.append(OperatorVerimlilik(
            operator_id=op.id,
            ad_soyad=op.kullanici.ad_soyad,
            sicil_no=op.sicil_no,
            tamamlanan_is=tamamlanan or 0,
            toplam_sure_dakika=toplam,
            ort_sure_dakika=ort,
        ))
    return result


@router.get("/makine-gecmis/{makine_id}", response_model=List[MakineGecmis])
def makine_gecmis(
    makine_id: int,
    baslangic: Optional[date] = Query(None),
    bitis: Optional[date] = Query(None),
    limit: int = Query(100, le=500),
    db: Session = Depends(get_db),
    _: Kullanici = Depends(get_current_user),
):
    from app.models.models import IsEmri, StokKarti
    q = (
        db.query(MakineIsEmri, IsEmri, StokKarti, Kullanici)
        .join(IsEmri, IsEmri.id == MakineIsEmri.is_emri_id)
        .join(StokKarti, StokKarti.id == IsEmri.stok_id)
        .join(Operator, Operator.id == MakineIsEmri.operator_id, isouter=True)
        .join(Kullanici, Kullanici.id == Operator.kullanici_id, isouter=True)
        .filter(MakineIsEmri.makine_id == makine_id)
    )
    if baslangic:
        q = q.filter(MakineIsEmri.baslangic_zamani >= datetime.combine(baslangic, datetime.min.time()))
    if bitis:
        q = q.filter(MakineIsEmri.baslangic_zamani <= datetime.combine(bitis, datetime.max.time()))
    rows = q.order_by(MakineIsEmri.baslangic_zamani.desc()).limit(limit).all()
    result = []
    for mie, ie, stok, kullanici in rows:
        sure = None
        if mie.baslangic_zamani and mie.bitis_zamani:
            sure = int((mie.bitis_zamani - mie.baslangic_zamani).total_seconds() / 60)
        result.append(MakineGecmis(
            makine_is_emri_id=mie.id, is_emri_no=ie.is_emri_no,
            stok_adi=stok.stok_adi, operator_adi=kullanici.ad_soyad if kullanici else "-",
            baslangic=mie.baslangic_zamani, bitis=mie.bitis_zamani,
            sure_dakika=sure, durum=mie.durum,
        ))
    return result


@router.get("/ariza-ozet")
def ariza_ozet(
    baslangic: Optional[date] = Query(None),
    bitis: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    _: Kullanici = Depends(get_current_user),
):
    q = (
        db.query(Makine.kod, Makine.ad,
                 func.count(Ariza.id).label("ariza_sayisi"))
        .join(Ariza, Ariza.makine_id == Makine.id)
    )
    if baslangic:
        q = q.filter(Ariza.baslangic >= datetime.combine(baslangic, datetime.min.time()))
    if bitis:
        q = q.filter(Ariza.baslangic <= datetime.combine(bitis, datetime.max.time()))
    q = q.group_by(Makine.id, Makine.kod, Makine.ad).order_by(func.count(Ariza.id).desc())

    result = []
    for kod, ad, sayisi in q.all():
        arizalar = db.query(Ariza).join(Makine).filter(Makine.kod == kod, Ariza.bitis != None).all()
        durus = sum(
            int((a.bitis - a.baslangic).total_seconds() / 60)
            for a in arizalar if a.bitis
        )
        result.append({"kod": kod, "ad": ad, "ariza_sayisi": sayisi, "toplam_durus_dakika": durus})
    return result


# ── EXCEL EXPORT ───────────────────────────────────────────

@router.get("/export/uretim-excel")
def uretim_excel(
    baslangic: Optional[date] = Query(None),
    bitis: Optional[date] = Query(None),
    makine_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    _: Kullanici = Depends(get_current_user),
):
    from app.models.models import IsEmri, StokKarti

    q = (
        db.query(MakineIsEmri, IsEmri, StokKarti, Kullanici, Makine)
        .join(IsEmri, IsEmri.id == MakineIsEmri.is_emri_id)
        .join(StokKarti, StokKarti.id == IsEmri.stok_id)
        .join(Makine, Makine.id == MakineIsEmri.makine_id)
        .join(Operator, Operator.id == MakineIsEmri.operator_id, isouter=True)
        .join(Kullanici, Kullanici.id == Operator.kullanici_id, isouter=True)
        .filter(MakineIsEmri.durum == "tamamlandi")
    )
    if makine_id:
        q = q.filter(MakineIsEmri.makine_id == makine_id)
    if baslangic:
        q = q.filter(MakineIsEmri.baslangic_zamani >= datetime.combine(baslangic, datetime.min.time()))
    if bitis:
        q = q.filter(MakineIsEmri.bitis_zamani <= datetime.combine(bitis, datetime.max.time()))
    rows = q.order_by(MakineIsEmri.baslangic_zamani.desc()).limit(1000).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Üretim Raporu"

    basliklar = ["İş Emri No", "Makine", "Ürün", "Operatör", "Başlangıç", "Bitiş", "Süre", "Durum"]
    fill, font, alignment = _excel_header_style()
    for col, baslik in enumerate(basliklar, 1):
        cell = ws.cell(row=1, column=col, value=baslik)
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment

    for row_idx, (mie, ie, stok, kullanici, makine) in enumerate(rows, 2):
        sure = None
        if mie.baslangic_zamani and mie.bitis_zamani:
            sure = int((mie.bitis_zamani - mie.baslangic_zamani).total_seconds() / 60)
        renk = "EBF5EB" if row_idx % 2 == 0 else "FFFFFF"
        fill2 = PatternFill("solid", fgColor=renk)
        satirlar = [
            ie.is_emri_no, makine.kod, stok.stok_adi,
            kullanici.ad_soyad if kullanici else "-",
            mie.baslangic_zamani.strftime("%d.%m.%Y %H:%M") if mie.baslangic_zamani else "-",
            mie.bitis_zamani.strftime("%d.%m.%Y %H:%M") if mie.bitis_zamani else "-",
            _dk_fmt(sure), mie.durum,
        ]
        for col, deger in enumerate(satirlar, 1):
            cell = ws.cell(row=row_idx, column=col, value=deger)
            cell.fill = fill2
            cell.alignment = Alignment(horizontal="center", vertical="center")

    genislikler = [18, 10, 25, 20, 18, 18, 10, 12]
    for col, g in enumerate(genislikler, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = g
    ws.row_dimensions[1].height = 25

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    tarih = datetime.now().strftime("%Y%m%d_%H%M")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=uretim_raporu_{tarih}.xlsx"}
    )


@router.get("/export/ariza-excel")
def ariza_excel(
    baslangic: Optional[date] = Query(None),
    bitis: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    _: Kullanici = Depends(get_current_user),
):
    q = (
        db.query(Ariza, Makine, Kullanici)
        .join(Makine, Makine.id == Ariza.makine_id)
        .join(Operator, Operator.id == Ariza.operator_id)
        .join(Kullanici, Kullanici.id == Operator.kullanici_id)
    )
    if baslangic:
        q = q.filter(Ariza.baslangic >= datetime.combine(baslangic, datetime.min.time()))
    if bitis:
        q = q.filter(Ariza.baslangic <= datetime.combine(bitis, datetime.max.time()))
    rows = q.order_by(Ariza.baslangic.desc()).limit(1000).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Arıza Raporu"
    basliklar = ["Makine", "Arıza Tipi", "Açıklama", "Bildiren", "Başlangıç", "Bitiş", "Süre", "Durum", "Çözüm"]
    fill, font, alignment = _excel_header_style()
    for col, baslik in enumerate(basliklar, 1):
        cell = ws.cell(row=1, column=col, value=baslik)
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment

    for row_idx, (ariza, makine, kullanici) in enumerate(rows, 2):
        sure = None
        if ariza.bitis:
            sure = int((ariza.bitis - ariza.baslangic).total_seconds() / 60)
        renk = "FFF3F3" if row_idx % 2 == 0 else "FFFFFF"
        fill2 = PatternFill("solid", fgColor=renk)
        satirlar = [
            makine.kod, ariza.ariza_tipi, ariza.aciklama or "-",
            kullanici.ad_soyad,
            ariza.baslangic.strftime("%d.%m.%Y %H:%M"),
            ariza.bitis.strftime("%d.%m.%Y %H:%M") if ariza.bitis else "-",
            _dk_fmt(sure), ariza.durum, ariza.cozum_aciklamasi or "-",
        ]
        for col, deger in enumerate(satirlar, 1):
            cell = ws.cell(row=row_idx, column=col, value=deger)
            cell.fill = fill2
            cell.alignment = Alignment(horizontal="center", vertical="center")

    genislikler = [10, 12, 30, 20, 18, 18, 10, 12, 30]
    for col, g in enumerate(genislikler, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = g

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    tarih = datetime.now().strftime("%Y%m%d_%H%M")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=ariza_raporu_{tarih}.xlsx"}
    )


@router.get("/export/uretim-pdf")
def uretim_pdf(
    baslangic: Optional[date] = Query(None),
    bitis: Optional[date] = Query(None),
    makine_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    _: Kullanici = Depends(get_current_user),
):
    from app.models.models import IsEmri, StokKarti

    q = (
        db.query(MakineIsEmri, IsEmri, StokKarti, Kullanici, Makine)
        .join(IsEmri, IsEmri.id == MakineIsEmri.is_emri_id)
        .join(StokKarti, StokKarti.id == IsEmri.stok_id)
        .join(Makine, Makine.id == MakineIsEmri.makine_id)
        .join(Operator, Operator.id == MakineIsEmri.operator_id, isouter=True)
        .join(Kullanici, Kullanici.id == Operator.kullanici_id, isouter=True)
        .filter(MakineIsEmri.durum == "tamamlandi")
    )
    if makine_id:
        q = q.filter(MakineIsEmri.makine_id == makine_id)
    if baslangic:
        q = q.filter(MakineIsEmri.baslangic_zamani >= datetime.combine(baslangic, datetime.min.time()))
    if bitis:
        q = q.filter(MakineIsEmri.bitis_zamani <= datetime.combine(bitis, datetime.max.time()))
    rows = q.order_by(MakineIsEmri.baslangic_zamani.desc()).limit(500).all()

    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4),
                            rightMargin=1*cm, leftMargin=1*cm,
                            topMargin=1.5*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    baslik_style = ParagraphStyle('baslik', parent=styles['Heading1'],
                                   fontSize=16, textColor=colors.HexColor('#1E3A5F'))
    alt_style = ParagraphStyle('alt', parent=styles['Normal'],
                                fontSize=10, textColor=colors.grey)

    elements = []
    elements.append(Paragraph("Üretim Raporu", baslik_style))
    elements.append(Paragraph(
        f"Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')} | Toplam: {len(rows)} kayıt",
        alt_style
    ))
    elements.append(Spacer(1, 0.5*cm))

    tablo_veri = [["İş Emri", "Makine", "Ürün", "Operatör", "Başlangıç", "Bitiş", "Süre"]]
    for mie, ie, stok, kullanici, makine in rows:
        sure = None
        if mie.baslangic_zamani and mie.bitis_zamani:
            sure = int((mie.bitis_zamani - mie.baslangic_zamani).total_seconds() / 60)
        adi = stok.stok_adi[:22] + "..." if len(stok.stok_adi) > 22 else stok.stok_adi
        tablo_veri.append([
            ie.is_emri_no, makine.kod, adi,
            (kullanici.ad_soyad if kullanici else "-")[:15],
            mie.baslangic_zamani.strftime("%d.%m.%Y %H:%M") if mie.baslangic_zamani else "-",
            mie.bitis_zamani.strftime("%d.%m.%Y %H:%M") if mie.bitis_zamani else "-",
            _dk_fmt(sure),
        ])

    tablo = Table(tablo_veri, repeatRows=1)
    tablo.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#F0F7F0'), colors.white]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CCCCCC')),
        ('ROWHEIGHT', (0,0), (-1,-1), 20),
    ]))
    elements.append(tablo)
    doc.build(elements)
    output.seek(0)

    tarih = datetime.now().strftime("%Y%m%d_%H%M")
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=uretim_raporu_{tarih}.pdf"}
    )